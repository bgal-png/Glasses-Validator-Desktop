"""Write the sheet out using the import template, so the header row keeps its
exact formatting (colours, fonts, wrap, row height, column widths).

Rather than trying to re-create the styling, we open `header_template.xlsx`,
leave row 1 untouched, and write the data underneath it. Data cells inherit
the number format of the template's first data row, which keeps things like
text-formatted barcodes intact.
"""
from __future__ import annotations
import os
import re
import sys
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from copy import copy

TEMPLATE_NAME = "header_template.xlsx"

ERROR_FILL = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFE08A", end_color="FFE08A", fill_type="solid")
ERROR_TYPES = {"empty_required", "empty_required_sun", "meta_format",
               "invalid_content", "duplicate"}


def resource_path(name: str) -> str:
    """Path to a bundled data file, working both in dev and in the PyInstaller exe."""
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, name)


def template_path() -> str | None:
    p = resource_path(TEMPLATE_NAME)
    return p if os.path.exists(p) else None


def _clean(v) -> str:
    return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""


def save_with_template(df, out_path, cell_issues=None, template=None):
    """Write df to out_path using the template's header row.

    cell_issues: optional {(row_idx, col_name): [issue, ...]} — when given, the
    flagged data cells are filled red/amber and carry a comment (annotated export).

    Returns (out_path, matched_columns, unmatched_columns).
    Raises FileNotFoundError if no template is available.
    """
    tpl = template or template_path()
    if not tpl:
        raise FileNotFoundError(f"{TEMPLATE_NAME} not found next to the app")

    wb = openpyxl.load_workbook(tpl)
    ws = wb.active

    # Map cleaned template header -> column index (1-based), keeping order
    tpl_cols = {}
    for c in range(1, ws.max_column + 1):
        h = _clean(ws.cell(row=1, column=c).value)
        if h and h not in tpl_cols:
            tpl_cols[h] = c

    # Remember the template's data-row formatting (e.g. barcode as text)
    proto = {}
    if ws.max_row >= 2:
        for c in range(1, ws.max_column + 1):
            src = ws.cell(row=2, column=c)
            proto[c] = (src.number_format, copy(src.font), copy(src.alignment))

    # Clear existing data rows, keep the header row intact
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    # Match our columns to template columns by cleaned header name
    matched, leftover = {}, []
    for col in df.columns:
        key = _clean(col)
        if key in tpl_cols:
            matched[col] = tpl_cols[key]
        else:
            leftover.append(col)

    # Fall back to position for unnamed columns: pandas calls a blank header
    # "Unnamed: N" (0-based), which lines up with template column N+1 when that
    # header is genuinely blank too.
    used = set(matched.values())
    unmatched = []
    for col in leftover:
        pos = df.columns.get_loc(col) + 1
        tpl_header_blank = (pos <= ws.max_column
                            and not _clean(ws.cell(row=1, column=pos).value))
        if tpl_header_blank and pos not in used:
            matched[col] = pos
            used.add(pos)
        else:
            unmatched.append(col)

    # Anything still unmatched is appended after the template's last column
    next_col = ws.max_column + 1
    for col in unmatched:
        ws.cell(row=1, column=next_col, value=str(col))
        matched[col] = next_col
        used.add(next_col)
        next_col += 1

    # Write the data
    for r in range(len(df)):
        excel_row = r + 2
        for col, c_idx in matched.items():
            v = df.iat[r, df.columns.get_loc(col)]
            v = "" if v is None or str(v) == "nan" else str(v)
            cell = ws.cell(row=excel_row, column=c_idx, value=v)
            fmt = proto.get(c_idx)
            if fmt:
                cell.number_format = fmt[0]
                cell.font = copy(fmt[1])
                cell.alignment = copy(fmt[2])
            if cell_issues:
                issues = cell_issues.get((r, col))
                if issues:
                    has_err = any(i["type"] in ERROR_TYPES for i in issues)
                    cell.fill = ERROR_FILL if has_err else WARNING_FILL
                    lines = []
                    for i in issues:
                        lines.append(f"- {i['message']}")
                        if i.get("expected"):
                            lines.append(f"    Expected: {i['expected']}")
                    cell.comment = Comment("\n".join(lines), "Validator")

    wb.save(out_path)
    return out_path, len(matched), unmatched
