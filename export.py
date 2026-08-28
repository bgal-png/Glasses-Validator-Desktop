"""Export the user's file back to .xlsx with flagged cells highlighted + commented."""
from __future__ import annotations
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

ERROR_FILL = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFE08A", end_color="FFE08A", fill_type="solid")
ERROR_TYPES = {"empty_required", "empty_required_sun", "meta_format", "invalid_content"}


def export_annotated(user_df, cell_issues, out_path):
    df = user_df.reset_index(drop=True)
    cols = list(df.columns)
    wb = Workbook()
    ws = wb.active
    ws.title = "Validated"

    # Header
    for c, name in enumerate(cols, start=1):
        ws.cell(row=1, column=c, value=name)
    ws.freeze_panes = "A2"

    # Data
    for r in range(len(df)):
        for c, name in enumerate(cols, start=1):
            v = df.iat[r, c - 1]
            cell = ws.cell(row=r + 2, column=c, value=("" if v is None else str(v)))
            issues = cell_issues.get((r, name))
            if issues:
                has_error = any(i["type"] in ERROR_TYPES for i in issues)
                cell.fill = ERROR_FILL if has_error else WARNING_FILL
                text = "\n".join(f"- {i['message']}" for i in issues)
                cell.comment = Comment(text, "Validator")

    # Reasonable column widths
    for c, name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(c)].width = min(max(10, len(str(name)) + 2), 40)

    wb.save(out_path)
    return out_path
