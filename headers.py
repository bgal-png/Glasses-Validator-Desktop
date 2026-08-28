"""Repair wrongly-named column headers using the import template as the
source of truth, so the validator's column mapping resolves fully.

Matching is layered, most-confident first:
  exact      — already identical to a template header (nothing to do)
  normalised — same letters/digits, differing punctuation/spacing/case
  no-id      — same descriptive part once the trailing "ID: nn" is dropped
  similar    — fuzzy match above a threshold
  position   — same column position, used only when the file and template
               have the same number of columns
"""
from __future__ import annotations
import difflib
import re

import openpyxl

import template

SIMILAR_THRESHOLD = 0.86


def _clean(v) -> str:
    return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""


def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", s.lower())


def _strip_id(s: str) -> str:
    """'Frame Colour ID: 26' -> 'framecolour'"""
    return _norm(re.sub(r"id\s*:?\s*\d+\s*$", "", s.strip(), flags=re.IGNORECASE))


def template_headers(path=None) -> list[str]:
    p = path or template.template_path()
    if not p:
        return []
    wb = openpyxl.load_workbook(p, read_only=True)
    ws = wb.active
    out = [_clean(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]
    wb.close()
    return out


def propose_fixes(file_cols, tpl_cols=None):
    """Suggest header renames.

    Returns (proposals, unmatched_template, unmatched_file) where each proposal
    is {"index", "current", "proposed", "method", "score"} for columns whose
    header should change.
    """
    tpl = tpl_cols if tpl_cols is not None else template_headers()
    if not tpl:
        return [], [], []

    file_cols = [str(c) for c in file_cols]
    same_length = len(file_cols) == len(tpl)

    tpl_by_norm, tpl_by_noid = {}, {}
    for t in tpl:
        if not t:
            continue
        tpl_by_norm.setdefault(_norm(t), t)
        tpl_by_noid.setdefault(_strip_id(t), t)

    taken = set()
    # Pass 0: exact (cleaned) matches are already correct — reserve them.
    exact = {}
    tpl_clean = {_clean(t) for t in tpl if t}
    for i, c in enumerate(file_cols):
        if _clean(c) in tpl_clean:
            exact[i] = _clean(c)
            taken.add(_clean(c))

    proposals = []
    for i, c in enumerate(file_cols):
        if i in exact:
            continue
        cn, cid = _norm(c), _strip_id(c)
        cand = method = None
        score = 1.0

        t = tpl_by_norm.get(cn)
        if t and t not in taken:
            cand, method = t, "normalised"

        if cand is None:
            t = tpl_by_noid.get(cid)
            if t and t not in taken and cid:
                cand, method = t, "no-id"

        if cand is None:
            pool = [t for t in tpl if t and t not in taken]
            best = difflib.get_close_matches(_clean(c), pool, n=1, cutoff=SIMILAR_THRESHOLD)
            if best:
                cand, method = best[0], "similar"
                score = difflib.SequenceMatcher(None, _clean(c), best[0]).ratio()

        # Position is only trusted when the header carries no real information
        # (blank / "Unnamed: n") or still resembles the template's header there,
        # otherwise it happily renames one column into an unrelated one.
        if cand is None and same_length and i < len(tpl) and tpl[i] and tpl[i] not in taken:
            cur = _clean(c)
            # "Uninformative" = carries no letters/digits at all ('', '-', '.',
            # '--') or is pandas' placeholder for a blank Excel header.
            blank = (not _norm(cur)) or re.fullmatch(r"unnamed:?\s*\d+", cur, re.IGNORECASE)
            ratio = difflib.SequenceMatcher(None, cur.lower(), tpl[i].lower()).ratio()
            if blank or ratio >= 0.5:
                cand, method, score = tpl[i], "position", ratio

        if cand and _clean(c) != cand:
            proposals.append({"index": i, "current": c, "proposed": cand,
                              "method": method, "score": score})
            taken.add(cand)

    unmatched_tpl = [t for t in tpl if t and t not in taken]
    proposed_idx = {p["index"] for p in proposals}
    unmatched_file = [c for i, c in enumerate(file_cols)
                      if i not in exact and i not in proposed_idx]
    return proposals, unmatched_tpl, unmatched_file


def apply_fixes(df, proposals):
    """Rename the given columns. Returns (new_df, applied_count)."""
    if not proposals:
        return df, 0
    cols = list(df.columns)
    for p in proposals:
        i = p["index"]
        if 0 <= i < len(cols):
            cols[i] = p["proposed"]
    df = df.copy()
    df.columns = cols
    return df, len(proposals)
